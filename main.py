#!/usr/bin/env python3
"""
Modbus Data Exporter GUI Application
Exports Modbus device data to CSV or Excel format
"""

import tkinter as tk
from tkinter import messagebox, filedialog, ttk
import threading
import time
import csv
import os
from datetime import datetime
import struct
import json
import weakref
from collections import OrderedDict
import hashlib
from functools import lru_cache
import asyncio
import concurrent.futures
import gc
import logging
import traceback
from typing import Optional, Dict, Any, List

# Version information
__version__ = "1.6.0"
__release_date__ = "2025-01-17"
__author__ = "Stefan Weidinger"

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler('modbus_exporter.log'),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger(__name__)

# Try to import modbus library
try:
    from pymodbus.client import ModbusTcpClient as ModbusClient
    MODBUS_AVAILABLE = True
except ImportError:
    try:
        from pymodbus.client.sync import ModbusTcpClient as ModbusClient
        MODBUS_AVAILABLE = True
    except ImportError:
        MODBUS_AVAILABLE = False

# Try to import openpyxl for Excel export
try:
    import openpyxl
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False

# Connection Pool Manager
class ConnectionPool:
    def __init__(self, max_connections=5):
        self.pool = {}
        self.max_connections = max_connections
        self.lock = threading.Lock()
    
    def get_connection(self, ip, port=502):
        """Get a connection from the pool or create a new one"""
        with self.lock:
            key = f"{ip}:{port}"
            if key in self.pool:
                client = self.pool[key]
                # Check if connection is still valid
                if hasattr(client, 'is_socket_open') and client.is_socket_open():
                    return client
                else:
                    # Remove invalid connection
                    del self.pool[key]
            
            # Create new connection if pool not full
            if len(self.pool) < self.max_connections:
                client = ModbusClient(ip, port=port)
                if client.connect():
                    # Store IP for caching purposes
                    client._cached_ip = ip
                    self.pool[key] = client
                    return client
            return None
    
    def close_all(self):
        """Close all connections in the pool"""
        with self.lock:
            for client in self.pool.values():
                try:
                    client.close()
                except:
                    pass
            self.pool.clear()

# Memory-efficient data cache
class DataCache:
    def __init__(self, max_size=100, ttl=300):  # 5 minutes TTL
        self.cache = OrderedDict()
        self.max_size = max_size
        self.ttl = ttl
        self.lock = threading.Lock()
    
    def _generate_key(self, ip, device_id, register, count):
        """Generate cache key"""
        return f"{ip}:{device_id}:{register}:{count}"
    
    def get(self, ip, device_id, register, count):
        """Get cached data if valid"""
        with self.lock:
            key = self._generate_key(ip, device_id, register, count)
            if key in self.cache:
                data, timestamp = self.cache[key]
                if time.time() - timestamp < self.ttl:
                    # Move to end (LRU)
                    self.cache.move_to_end(key)
                    return data
                else:
                    del self.cache[key]
            return None
    
    def set(self, ip, device_id, register, count, data):
        """Cache data with timestamp"""
        with self.lock:
            key = self._generate_key(ip, device_id, register, count)
            # Remove oldest if cache is full
            if len(self.cache) >= self.max_size:
                self.cache.popitem(last=False)
            self.cache[key] = (data, time.time())
    
    def clear(self):
        """Clear all cached data"""
        with self.lock:
            self.cache.clear()

# Global instances
connection_pool = ConnectionPool()
data_cache = DataCache()

# Memory management utilities
class MemoryManager:
    def __init__(self):
        self.weak_refs = set()
    
    def add_reference(self, obj):
        """Add weak reference to track object"""
        self.weak_refs.add(weakref.ref(obj))
    
    def cleanup(self):
        """Clean up dead references"""
        self.weak_refs = {ref for ref in self.weak_refs if ref() is not None}
    
    def get_memory_usage(self):
        """Get current memory usage statistics"""
        import psutil
        import os
        try:
            process = psutil.Process(os.getpid())
            return process.memory_info().rss / 1024 / 1024  # MB
        except:
            return 0

memory_manager = MemoryManager()

# Enhanced error handling system
class ModbusError(Exception):
    """Custom exception for Modbus operations"""
    def __init__(self, message, error_code=None, details=None):
        super().__init__(message)
        self.error_code = error_code
        self.details = details or {}
        self.timestamp = datetime.now()
        logger.error(f"ModbusError: {message} (Code: {error_code})")

class ConnectionError(ModbusError):
    """Exception for connection failures"""
    def __init__(self, message, ip=None, port=502):
        super().__init__(message, "CONNECTION_FAILED", {"ip": ip, "port": port})
        self.ip = ip
        self.port = port

class DataValidationError(ModbusError):
    """Exception for data validation failures"""
    def __init__(self, message, field=None, value=None):
        super().__init__(message, "DATA_VALIDATION_FAILED", {"field": field, "value": value})
        self.field = field
        self.value = value

class AsyncOperationManager:
    """Manages asynchronous operations with proper error handling"""
    def __init__(self, max_workers=4):
        self.executor = concurrent.futures.ThreadPoolExecutor(max_workers=max_workers)
        self.active_futures = set()
        self.logger = logging.getLogger(__name__ + '.AsyncOperationManager')
    
    def submit_task(self, func, *args, **kwargs):
        """Submit a task for asynchronous execution"""
        future = self.executor.submit(self._wrapped_execution, func, *args, **kwargs)
        self.active_futures.add(future)
        future.add_done_callback(self._task_completed)
        return future
    
    def _wrapped_execution(self, func, *args, **kwargs):
        """Wrapped execution with error handling"""
        try:
            return func(*args, **kwargs)
        except Exception as e:
            self.logger.error(f"Async task failed: {e}")
            self.logger.debug(f"Traceback: {traceback.format_exc()}")
            raise
    
    def _task_completed(self, future):
        """Callback when task is completed"""
        self.active_futures.discard(future)
        if future.exception():
            self.logger.error(f"Task completed with exception: {future.exception()}")
    
    def shutdown(self, wait=True):
        """Shutdown the executor"""
        self.executor.shutdown(wait=wait)
        self.active_futures.clear()

# Global async manager
async_manager = AsyncOperationManager()

# Optimized decode functions with caching
@lru_cache(maxsize=1000)
def decode_ascii_tuple(registers_tuple):
    """Cached ASCII decode function for tuple input"""
    registers = list(registers_tuple)
    return "".join(
        chr((reg >> 8) & 0xFF) + chr(reg & 0xFF) for reg in registers
    ).split("\x00")[0].strip()

# Wrapper for tuple conversion
def decode_ascii_cached(registers):
    """Convert list to tuple for caching"""
    if registers:
        return decode_ascii_tuple(tuple(registers))
    return ""

def decode_float32(registers):
    """Decode Float32 value from two Modbus registers."""
    if registers and len(registers) == 2:
        combined = (registers[0] << 16) | registers[1]
        return struct.unpack('!f', struct.pack('!I', combined))[0]
    return None

# Optimized read_registers function with caching
def read_registers(client, device_id, address, count, log_widget=None):
    # Check cache first
    ip = getattr(client, '_cached_ip', 'unknown')
    cached_data = data_cache.get(ip, device_id, address, count)
    if cached_data is not None:
        return cached_data
    
    try:
        # Try newer parameter style first
        try:
            result = client.read_holding_registers(address, count=count, slave=device_id)
        except TypeError:
            # Fall back to older parameter style
            result = client.read_holding_registers(address, count, unit=device_id)
        
        if result.isError():
            raise Exception(f"Modbus-Fehler: {result}")
        
        # Cache the result
        data_cache.set(ip, device_id, address, count, result.registers)
        return result.registers
    except Exception as e:
        if log_widget:
            log_widget.log_message(f"⚠ Fehler beim Lesen der Register {address}: {e}")
        return None

def get_signal_quality(lqi, per):
    """Calculate signal quality level based on LQI and PER values
    Based on Schneider Electric EcoStruxure Panel Server documentation
    
    Signal quality matrix:
                    LQI < 30    30 ≤ LQI < 60    60 ≤ LQI
    PER > 30%         Weak         Weak          Fair
    10% < PER ≤ 30%   Weak         Fair          Good
    PER ≤ 10%         Fair         Good          Excellent
    """
    # Handle invalid or missing values
    if lqi is None or per is None:
        return "Unknown"
    
    try:
        lqi_value = float(lqi)
        per_value = float(per)
        
        # Handle NaN values
        if str(lqi_value).lower() == 'nan' or str(per_value).lower() == 'nan':
            return "Unknown"
        
        # Apply the signal quality matrix
        if per_value > 30:
            if lqi_value < 30:
                return "Weak"
            elif lqi_value < 60:
                return "Weak"
            else:  # lqi_value >= 60
                return "Fair"
        elif per_value > 10:
            if lqi_value < 30:
                return "Weak"
            elif lqi_value < 60:
                return "Fair"
            else:  # lqi_value >= 60
                return "Good"
        else:  # per_value <= 10
            if lqi_value < 30:
                return "Fair"
            elif lqi_value < 60:
                return "Good"
            else:  # lqi_value >= 60
                return "Excellent"
                
    except (ValueError, TypeError):
        return "Unknown"

def decode_heattag_alarm_type(value):
    """Decode HeatTag alarm type value to human-readable string"""
    if value is None or value == "N/A":
        return "N/A"
    
    try:
        val = int(value)
        if val == 0:
            return "No alarm"
        elif 1 <= val <= 15:
            return "Low level alarm"
        elif 16 <= val <= 93:
            return "Medium level alarm"
        elif val == 99:
            return "Test alarm"
        elif 94 <= val <= 190:
            return "High level alarm"
        else:
            return f"Unknown ({val})"
    except (ValueError, TypeError):
        return f"Invalid ({value})"

def decode_heattag_alarm_level(value):
    """Decode HeatTag alarm level value to human-readable string"""
    if value is None or value == "N/A":
        return "N/A"
    
    try:
        val = int(value)
        if val == 0:
            return "No alarm"
        elif val == 1:
            return "Low level alarm"
        elif val == 2:
            return "Medium level alarm"
        elif val == 3:
            return "High level alarm"
        else:
            return f"Unknown ({val})"
    except (ValueError, TypeError):
        return f"Invalid ({value})"

def decode_heattag_operation_mode(value):
    """Decode HeatTag operation mode value to human-readable string"""
    if value is None or value == "N/A":
        return "N/A"
    
    try:
        val = int(value)
        if val == 0:
            return "Test mode (0-30 min after power on)"
        elif val == 1:
            return "Auto-learning mode (30 min-8 hrs after power on)"
        elif val == 2:
            return "Normal operation mode (>8 hrs after power on)"
        else:
            return f"Unknown ({val})"
    except (ValueError, TypeError):
        return f"Invalid ({value})"

def decode_communication_status(value):
    """Decode Communication Status value to human-readable string"""
    if value is None or value == "N/A":
        return "N/A"
    
    try:
        val = int(value)
        if val == 0:
            return "Com. loss"
        elif val == 1:
            return "OK"
        else:
            return f"Unknown ({val})"
    except (ValueError, TypeError):
        return f"Invalid ({value})"

def decode_rf_communication_validity(value):
    """Decode RF Communication Validity value to human-readable string"""
    if value is None or value == "N/A":
        return "N/A"
    
    try:
        val = int(value)
        if val == 0:
            return "Invalid"
        elif val == 1:
            return "Valid"
        else:
            return f"Unknown ({val})"
    except (ValueError, TypeError):
        return f"Invalid ({value})"

def read_enhanced_diagnostics(client, device_id, device_type, log_widget=None):
    """Read enhanced diagnostics for TH110, CL110, and HeatTag devices"""
    diagnostics = {}
    # Define common registers for all device types
    enhanced_registers = [
        (31144, 1, "RF Communication Validity", "BITMAP"),
        (31145, 1, "Communication Status", "BITMAP"),
        (31151, 2, "Gateway PER", "Float32"),
        (31153, 2, "RSSI", "Float32"),
        (31155, 1, "LQI", "UINT16"),
        (31156, 2, "PER Max", "Float32"),
        (31158, 2, "RSSI Min", "Float32"),
        (31160, 1, "LQI Min", "UINT16")
    ]
    
    # Add device-specific registers
    if device_type == "CL110":
        enhanced_registers.append((3315, 2, "Battery Voltage", "Float32"))
    elif device_type == "HeatTag":
        # HeatTag specific registers - only for HeatTag devices
        enhanced_registers.extend([
            (3321, 1, "HeatTag Alarm Type", "UINT16"),
            (3322, 1, "HeatTag Alarm Level", "UINT16"),
            (31175, 1, "HeatTag Operation Mode", "UINT16")
        ])
    
    for addr, count, field_name, field_type in enhanced_registers:
        regs = read_registers(client, device_id, addr, count, log_widget)
        if regs:
            value = None
            if field_type == "Float32":
                value = round(decode_float32(regs), 2)
            elif field_type == "UINT16":
                value = regs[0] if regs else None
            elif field_type == "BITMAP":
                value = regs[0] if regs else None
            # Add more type handlers if needed
            
            diagnostics[field_name] = value
            if log_widget:
                log_widget.log_message(f"  ✓ {field_name}: {value}")
        else:
            diagnostics[field_name] = "N/A"
            if log_widget:
                log_widget.log_message(f"  ⚠ {field_name}: Error reading")
    
    # Calculate Signal Quality based on LQI and PER
    lqi_value = diagnostics.get("LQI")
    per_value = diagnostics.get("Gateway PER")
    signal_quality = get_signal_quality(lqi_value, per_value)
    diagnostics["Signal Quality"] = signal_quality
    if log_widget:
        log_widget.log_message(f"  ✓ Signal Quality: {signal_quality}")
    
    return diagnostics

# Original get_device_ids function
def get_device_ids(client, log_widget=None):
    base = 504
    step = 5
    max_devices = 100
    device_ids = []

    if log_widget:
        log_widget.log_message("→ Suche DeviceIDs in alternativen Registern (504, 509, 514, ...)")
    
    for i in range(max_devices):
        addr = base + (i * step)
        result = read_registers(client, 255, addr, 1, log_widget)
        if result and result[0] not in (0, 0xFFFF):
            device_id = result[0]
            if log_widget:
                log_widget.log_message(f"✓ Reg {addr}: DeviceID {device_id}")
            device_ids.append(device_id)
        else:
            if log_widget:
                log_widget.log_message(f"- Kein gültiger DeviceID-Wert in Register {addr}")
    return device_ids

# Optimized collect_data function with connection pooling
def collect_data(ip, log_widget=None):
    # Use connection pool
    client = connection_pool.get_connection(ip)
    if not client:
        if log_widget:
            log_widget.log_message("❌ Verbindung fehlgeschlagen.")
        return None

    if log_widget:
        log_widget.log_message("✓ Verbindung erfolgreich hergestellt.")
    
    device_ids = get_device_ids(client, log_widget)
    if not device_ids:
        if log_widget:
            log_widget.log_message("⚠ Keine gültigen DeviceIDs gefunden.")
        client.close()
        return None

    data = []
    for idx, device_id in enumerate(device_ids, start=1):
        if log_widget:
            log_widget.log_message(f"[{idx}/{len(device_ids)}] Verarbeite Device ID {device_id}")
        
        device_data = {
            "DeviceID": device_id,
            "DeviceType": "",
            "RFID": "",
            "SerialNumber": "",
        }

        # Commercial Reference → 31060
        ref_regs = read_registers(client, device_id, 31060, 16, log_widget)
        ref = decode_ascii_cached(ref_regs) if ref_regs else ""
        if log_widget:
            log_widget.log_message(f"→ Device {device_id} hat Commercial Reference: {ref}")

        device_type = ""
        if ref == "EMS59443":
            device_type = "CL110"
        elif ref == "EMS59440":
            device_type = "TH110"
        elif ref == "SMT10020":
            device_type = "HeatTag"
        else:
            device_type = "Unknown"
        device_data["DeviceType"] = device_type

        # RFID → 31026 (6 Register, hex)
        rfid_regs = read_registers(client, device_id, 31026, 6, log_widget)
        if rfid_regs:
            if log_widget:
                log_widget.log_message(f"  📦 RFID (Reg 31026, 6): {rfid_regs}")
            hex_str = "".join(f"{reg:04X}" for reg in rfid_regs if reg > 0)
            device_data["RFID"] = hex_str[:8]
        else:
            if log_widget:
                log_widget.log_message("  ⚠ RFID: Fehler beim Lesen")

        # Serial Number → 31088 (10 Register, ASCII)
        sn_regs = read_registers(client, device_id, 31088, 10, log_widget)
        if sn_regs:
            sn = decode_ascii_cached(sn_regs)
            if log_widget:
                log_widget.log_message(f"  📦 SerialNumber (Reg 31088, 10): {sn_regs}")
                log_widget.log_message(f"  ✓ SerialNumber: {sn}")
            device_data["SerialNumber"] = sn
        else:
            if log_widget:
                log_widget.log_message("  ⚠ SerialNumber: Fehler beim Lesen")
        
        # Device Name → 31000 (10 Register, ASCII)
        device_name_regs = read_registers(client, device_id, 31000, 10, log_widget)
        if device_name_regs:
            device_name = decode_ascii_cached(device_name_regs)
            if log_widget:
                log_widget.log_message(f"  📦 DeviceName (Reg 31000, 10): {device_name_regs}")
                log_widget.log_message(f"  ✓ DeviceName: {device_name}")
            device_data["DeviceName"] = device_name
        else:
            if log_widget:
                log_widget.log_message("  ⚠ DeviceName: Fehler beim Lesen")
            device_data["DeviceName"] = ""
        
        # Device Label → 31010 (3 Register, ASCII)
        device_label_regs = read_registers(client, device_id, 31010, 3, log_widget)
        if device_label_regs:
            device_label = decode_ascii_cached(device_label_regs)
            if log_widget:
                log_widget.log_message(f"  📦 DeviceLabel (Reg 31010, 3): {device_label_regs}")
                log_widget.log_message(f"  ✓ DeviceLabel: {device_label}")
            device_data["DeviceLabel"] = device_label
        else:
            if log_widget:
                log_widget.log_message("  ⚠ DeviceLabel: Fehler beim Lesen")
            device_data["DeviceLabel"] = ""

        # Enhanced Diagnostics if enabled
        if hasattr(log_widget, 'enhanced_diagnostics_var') and log_widget.enhanced_diagnostics_var.get() and device_type in ["TH110", "CL110", "HeatTag"]:
            enhanced_diagnostics = read_enhanced_diagnostics(client, device_id, device_type, log_widget)
            device_data["EnhancedDiagnostics"] = enhanced_diagnostics
            if log_widget:
                log_widget.log_message(f"→ Enhanced Diagnostics for {device_type}: {enhanced_diagnostics}")
        else:
            device_data["EnhancedDiagnostics"] = {}

        # Product Model (nur Debug) → 31106
        pm_regs = read_registers(client, device_id, 31106, 8, log_widget)
        if pm_regs:
            pm = decode_ascii_cached(pm_regs)
            if log_widget:
                log_widget.log_message(f"  📦 ProductModel (Reg 31106, 8): {pm_regs}")
                log_widget.log_message(f"  ✓ ProductModel: {pm}")

        data.append(device_data)

    client.close()
    return data

class ModbusExporterGUI:
    def __init__(self, root):
        self.root = root
        self.root.title(f"Modbus Data Exporter v{__version__}")
        self.root.geometry("1600x900")
        self.root.configure(bg='#282a36')  # Modern dark theme background
        
        # Modern resizable window with improved constraints
        self.root.resizable(True, True)
        self.root.minsize(1400, 700)
        
        # Set window icon if available (optional)
        try:
            # You can add an icon here if you have one
            pass
        except:
            pass
        
        
        # Variables
        self.is_running = False
        self.export_thread = None
        self.csv_var = tk.BooleanVar(value=True)
        self.excel_var = tk.BooleanVar(value=EXCEL_AVAILABLE)
        self.enhanced_diagnostics_var = tk.BooleanVar(value=False)
        self.sensor_pairing_var = tk.BooleanVar(value=False)
        
        # Live diagnostics variables
        self.live_diagnostics_enabled = False
        self.live_diagnostics_thread = None
        self.live_data_tree_columns = ["DeviceID", "DeviceType", "RFID", "SerialNumber", "DeviceName", "RFCommunication", "CommStatus", "SignalQuality", "RSSI", "LQI", "GatewayPER", "Battery"]
        self.last_connection_test = False
        self.last_live_update = "Never"
        
        # Column visibility variables (only for optional columns)
        self.always_visible_columns = ["DeviceID", "DeviceType", "RFID", "SerialNumber"]
        self.optional_columns = [col for col in self.live_data_tree_columns if col not in self.always_visible_columns]
        self.column_visibility = {}
        for col in self.optional_columns:
            self.column_visibility[col] = tk.BooleanVar(value=True)
        
        # Setup GUI
        self.setup_gui()
        
        # Handle window closing
        self.root.protocol("WM_DELETE_WINDOW", self.on_closing)
        
        # Bind ESC key to exit
        self.root.bind('<Escape>', lambda e: self.on_closing())
        
        # Make window focusable
        self.root.focus_force()
    
    def create_tooltip(self, widget, text):
        """Create a tooltip for a widget"""
        def on_enter(event):
            tooltip = tk.Toplevel()
            tooltip.wm_overrideredirect(True)
            tooltip.wm_geometry(f"+{event.x_root+10}+{event.y_root+10}")
            tooltip.configure(bg='#ffffe0', bd=1, relief='solid')
            label = tk.Label(tooltip, text=text, bg='#ffffe0', fg='#000000', 
                           font=('Arial', 9), pady=2, padx=5)
            label.pack()
            widget.tooltip = tooltip
        
        def on_leave(event):
            if hasattr(widget, 'tooltip'):
                widget.tooltip.destroy()
                del widget.tooltip
        
        widget.bind('<Enter>', on_enter)
        widget.bind('<Leave>', on_leave)

    def setup_gui(self):
        """Setup the main GUI elements"""
        
        # Header frame with modern gradient-like effect
        header_frame = tk.Frame(self.root, bg='#6272a4', relief='flat')
        header_frame.pack(fill='x', pady=(0, 10))
        
        # Title with professional styling
        title_label = tk.Label(header_frame, text="Modbus Data Exporter Professional", 
                              font=("Helvetica Neue", 24, "bold"), 
                              bg='#6272a4', fg='#f8f8f2')
        title_label.pack(side='left', padx=(30, 0), pady=15)
        
        # Version info with modern styling
        version_frame = tk.Frame(header_frame, bg='#44475a', relief='flat', bd=0)
        version_frame.pack(side='right', padx=(0, 30), pady=10)
        
        version_label = tk.Label(version_frame, text=f"v{__version__}", 
                                font=("Helvetica Neue", 12, "bold"), 
                                bg='#44475a', fg='#50fa7b')
        version_label.pack(padx=15, pady=8)
        
        # Create tooltip for version info
        self.create_tooltip(version_label, f"Released: {__release_date__}\nBy: {__author__}")
        
        # Main container with modern styling
        main_container = tk.Frame(self.root, bg='#282a36')
        main_container.pack(fill='both', expand=True, padx=20, pady=10)
        
        # Configure column weights for proportional sizing - make left panel smaller
        main_container.grid_columnconfigure(0, weight=1, minsize=250)  # Left column smaller
        main_container.grid_columnconfigure(1, weight=7, minsize=1200)  # Right column much larger
        main_container.grid_rowconfigure(0, weight=1)      # Single row fills height
        
        # Left column with explicit width constraint - reduced size
        left_column = tk.Frame(main_container, bg='#282a36', width=280)
        left_column.grid(row=0, column=0, sticky='nsew', padx=(0, 10))
        left_column.grid_propagate(False)  # Don't let contents resize the frame
        
        # Right column takes remaining space
        right_column = tk.Frame(main_container, bg='#282a36')
        right_column.grid(row=0, column=1, sticky='nsew', padx=(10, 0))
        
        # Adjust weight of row containing checkboxes to reduce resizing
        right_column.grid_rowconfigure(1, weight=1)  # Ensure stability of Live View
        
        # === LEFT COLUMN CONTENT ===
        
        # IP Address Section with modern card design
        ip_frame = tk.Frame(left_column, bg='#44475a', relief='flat', bd=0)
        ip_frame.pack(fill='x', pady=(0, 15))
        
        ip_label = tk.Label(ip_frame, text="Device IP Address", 
                           font=("Helvetica Neue", 14, "bold"), 
                           bg='#44475a', fg='#f8f8f2')
        ip_label.pack(pady=(15, 10))
        
        # Modern IP Entry with professional styling
        self.ip_entry = tk.Entry(ip_frame, width=25, font=("Helvetica Neue", 12),
                                relief='flat', bd=0, bg='#6272a4', fg='#f8f8f2',
                                highlightbackground='#44475a',
                                highlightcolor='#50fa7b', highlightthickness=2,
                                insertbackground='#f8f8f2')
        self.ip_entry.pack(pady=(0, 15), ipady=8)
        self.ip_entry.config(state='normal')  # Make sure the entry is enabled
        self.ip_entry.focus_set()  # Set focus to the IP entry
        self.ip_entry.insert(0, "10.0.1.110")  # Default value
        
        # Test IP Button with modern design
        test_frame = tk.Frame(left_column, bg='#282a36')
        test_frame.pack(fill='x', pady=(0, 15))
        
        # Professional Test Button
# Button with rounded corners and darker text
        test_btn = tk.Button(test_frame, text="Test Connection", 
                            command=self.test_ip,
                            font=("Helvetica Neue", 12, "bold"),
                            bg='#8be9fd', fg='#2d2d2d',
                            relief='flat', bd=0,
                            activebackground='#50fa7b',
                            activeforeground='#2d2d2d',
                            highlightthickness=0, highlightbackground='#2d2d2d',
                            highlightcolor='#2d2d2d')
        test_btn.config(borderwidth=2, overrelief='solid')
        test_btn.pack(fill='x', padx=0, pady=0, ipady=10)
        
        # Export Options with modern card design
        export_frame = tk.Frame(left_column, bg='#44475a', relief='flat', bd=0)
        export_frame.pack(fill='x', pady=(0, 15))
        
        export_label = tk.Label(export_frame, text="Export Options", 
                               font=("Helvetica Neue", 14, "bold"), 
                               bg='#44475a', fg='#f8f8f2')
        export_label.pack(pady=(15, 10))
        
        # Modern checkboxes with professional styling
        csv_cb = tk.Checkbutton(export_frame, text="Export to CSV",
                               variable=self.csv_var, font=("Helvetica Neue", 11),
                               bg='#44475a', fg='#f8f8f2', activeforeground='#50fa7b',
                               activebackground='#44475a', selectcolor='#6272a4')
        csv_cb.pack(pady=5, padx=15, anchor='w')
        
        # Excel Checkbox with modern styling
        excel_text = "Export to Excel" if EXCEL_AVAILABLE else "Export to Excel (openpyxl not installed)"
        excel_cb = tk.Checkbutton(export_frame, text=excel_text,
                                 variable=self.excel_var, font=("Helvetica Neue", 11),
                                 bg='#44475a', fg='#f8f8f2', activeforeground='#50fa7b',
                                 activebackground='#44475a', selectcolor='#6272a4',
                                 disabledforeground='#6272a4',
                                 state='normal' if EXCEL_AVAILABLE else 'disabled')
        excel_cb.pack(pady=5, padx=15, anchor='w')
        
        # Enhanced Diagnostics Checkbox
        enhanced_diag_cb = tk.Checkbutton(export_frame, text="Enable Enhanced Diagnostics",
                                       variable=self.enhanced_diagnostics_var, font=("Helvetica Neue", 11),
                                       bg='#44475a', fg='#f8f8f2', activeforeground='#50fa7b',
                                       activebackground='#44475a', selectcolor='#6272a4')
        enhanced_diag_cb.pack(pady=5, padx=15, anchor='w')
        
        # Sensor Pairing Sheet Checkbox
        sensor_pairing_cb = tk.Checkbutton(export_frame, text="Generate Sensor Pairing Sheet",
                                         variable=self.sensor_pairing_var, font=("Helvetica Neue", 11),
                                         bg='#44475a', fg='#f8f8f2', activeforeground='#50fa7b',
                                         activebackground='#44475a', selectcolor='#6272a4')
        sensor_pairing_cb.pack(pady=(5, 15), padx=15, anchor='w')

        # Control Buttons with modern design
        button_frame = tk.Frame(left_column, bg='#282a36')
        button_frame.pack(pady=15)
        
# Button with rounded corners and darker text
        self.start_btn = tk.Button(button_frame, text="START EXPORT",
                                  command=self.start_export,
                                  bg='#50fa7b', fg='#2d2d2d',
                                  font=("Helvetica Neue", 12, "bold"),
                                  width=20, height=2, relief='flat', bd=0,
                                  activebackground='#8be9fd',
                                  activeforeground='#2d2d2d',
                                  highlightthickness=0, highlightbackground='#2d2d2d',
                                  highlightcolor='#2d2d2d')
        self.start_btn.config(borderwidth=2, overrelief='solid')
        self.start_btn.pack(pady=5)
        
        
        # Status Section with modern design
        status_frame = tk.Frame(left_column, bg='#44475a', relief='flat', bd=0)
        status_frame.pack(fill='x', pady=(0, 15))
        
        status_title = tk.Label(status_frame, text="Status", 
                               font=("Helvetica Neue", 14, "bold"), 
                               bg='#44475a', fg='#f8f8f2')
        status_title.pack(pady=(15, 5))
        
        self.status_label = tk.Label(status_frame, text="Ready", 
                                    font=("Helvetica Neue", 12), 
                                    bg='#44475a', fg='#50fa7b')
        self.status_label.pack(pady=(0, 15))
        
        # Activity Log Button
        log_button_frame = tk.Frame(left_column, bg='#282a36')
        log_button_frame.pack(pady=(15, 0))
        
        # Activity Log Button with modern design
        self.log_btn = tk.Button(log_button_frame, text="View Activity Log",
                                command=self.open_log_window,
                                bg='#6272a4', fg='#2d2d2d',
                                font=("Helvetica Neue", 12, "bold"),
                                width=20, height=2, relief='flat', bd=0,
                                activebackground='#44475a',
                                activeforeground='#2d2d2d',
                                highlightthickness=0, highlightbackground='#2d2d2d',
                                highlightcolor='#2d2d2d')
        self.log_btn.config(borderwidth=2, overrelief='solid')
        self.log_btn.pack(pady=5)
        
        # Initialize log window reference
        self.log_window = None
        self.log_text = None  # Will be created when log window opens
        
        # === RIGHT COLUMN CONTENT - Live Diagnostics View ===
        
        # Live Diagnostics Header with modern styling
        live_diag_header = tk.Frame(right_column, bg='#44475a', relief='flat', bd=0)
        live_diag_header.pack(fill='x', pady=(0, 15))
        
        # Title row with status icon and last update in one line
        title_row = tk.Frame(live_diag_header, bg='#44475a')
        title_row.pack(fill='x', pady=(15, 10))
        
        # Status icon (left side) - round colored circle
        self.status_icon = tk.Label(title_row, text="●", 
                                   font=("Helvetica Neue", 20), 
                                   bg='#44475a', fg='#ff5555')  # Red by default
        self.status_icon.pack(side='left', padx=(10, 0))
        
        # Last Update Timestamp (right side)
        self.last_update_label = tk.Label(title_row, text="Last Update: Never", 
                                         font=("Helvetica Neue", 10), 
                                         bg='#44475a', fg='#6272a4')
        self.last_update_label.pack(side='right')
        
        # Live Diagnostics title (centered)
        live_diag_title = tk.Label(title_row, text="Live Diagnostics View", 
                                  font=("Helvetica Neue", 16, "bold"), 
                                  bg='#44475a', fg='#f8f8f2')
        live_diag_title.pack(expand=True)  # This centers the title
        
        # Live Diagnostics Button with modern styling
# Button with rounded corners and darker text
        self.live_diag_btn = tk.Button(live_diag_header, text="Start Live Diagnostics",
                                      command=self.toggle_live_diagnostics,
                                      bg='#50fa7b', fg='#2d2d2d',
                                      font=("Helvetica Neue", 12, "bold"),
                                      width=20, height=2, relief='flat', bd=0,
                                      activebackground='#8be9fd',
                                      activeforeground='#2d2d2d',
                                      state='disabled',
                                      highlightthickness=0, highlightbackground='#2d2d2d',
                                      highlightcolor='#2d2d2d')
        self.live_diag_btn.config(borderwidth=2, overrelief='solid')
        self.live_diag_btn.pack(pady=10)
        
        # Column Visibility Controls with modern styling
        columns_frame = tk.Frame(live_diag_header, bg='#44475a')
        columns_frame.pack(pady=10, fill='x')  # Ensure consistent width for stability
        
        columns_label = tk.Label(columns_frame, text="Visible Columns:", 
                                font=("Helvetica Neue", 11, "bold"), 
                                bg='#44475a', fg='#f8f8f2')
        columns_label.pack(pady=2)
        
        # Create a frame for column checkboxes in a grid layout
        column_checkboxes_frame = tk.Frame(columns_frame, bg='#44475a')
        column_checkboxes_frame.pack(pady=5)
        
        # Column display names (only for optional columns)
        column_display_names = {
            "DeviceName": "Name",
            "RFCommunication": "RF Com.",
            "CommStatus": "Com. Status",
            "SignalQuality": "Signal Quality",
            "RSSI": "RSSI",
            "LQI": "LQI",
            "GatewayPER": "Gateway PER",
            "Battery": "Battery"
        }
        
        # Create checkboxes only for optional columns in a single line with modern styling
        for col in self.optional_columns:
            cb = tk.Checkbutton(column_checkboxes_frame, 
                               text=column_display_names.get(col, col),
                               variable=self.column_visibility[col],
                               command=self.update_column_visibility,
                               font=("Helvetica Neue", 9), 
                               bg='#44475a', fg='#f8f8f2',
                               activeforeground='#50fa7b', 
                               activebackground='#44475a',
                               selectcolor='#6272a4')
            cb.pack(side='left', padx=8, pady=2)  # Single line layout
        
        # Live Diagnostics Data Frame with modern styling
        live_data_frame = tk.Frame(right_column, bg='#44475a', relief='flat', bd=0)
        live_data_frame.pack(fill='both', expand=True, pady=0)
        
        # Set a fixed width for the live data frame to prevent resizing
        live_data_frame.pack_propagate(False)
        
        # Live data display tree with improved styling
        self.live_data_tree = ttk.Treeview(live_data_frame, columns=self.live_data_tree_columns, show='headings')
        
        # Define column headings and widths
        column_config = {
            "DeviceID": {"text": "ID", "width": 60, "anchor": tk.CENTER},
            "DeviceType": {"text": "Type", "width": 140, "anchor": tk.CENTER},  # Increased width for better readability
            "RFID": {"text": "RFID", "width": 120, "anchor": tk.CENTER},  # Increased width for 8 chars
            "SerialNumber": {"text": "Serial Number", "width": 160, "anchor": tk.CENTER},
            "DeviceName": {"text": "Name", "width": 120, "anchor": tk.W},
            "RFCommunication": {"text": "RF Com.", "width": 80, "anchor": tk.CENTER},
            "CommStatus": {"text": "Com. Status", "width": 80, "anchor": tk.CENTER},
            "SignalQuality": {"text": "Signal Quality", "width": 100, "anchor": tk.CENTER},
            "RSSI": {"text": "RSSI (dBm)", "width": 80, "anchor": tk.CENTER},
            "LQI": {"text": "LQI", "width": 60, "anchor": tk.CENTER},
            "GatewayPER": {"text": "Gateway PER", "width": 80, "anchor": tk.CENTER},
            "Battery": {"text": "Battery (V)", "width": 80, "anchor": tk.CENTER}
        }
        
        for col in self.live_data_tree_columns:
            config = column_config.get(col, {"text": col, "width": 100, "anchor": tk.CENTER})
            self.live_data_tree.heading(col, text=config["text"])
            self.live_data_tree.column(col, width=config["width"], anchor=config["anchor"], minwidth=50)
        
        # Configure tree view styling with modern color scheme
        style = ttk.Style()
        style.theme_use('clam')
        style.configure('Treeview', background='#f8f8f2', foreground='#282a36', fieldbackground='#f8f8f2')
        style.configure('Treeview.Heading', background='#6272a4', foreground='#f8f8f2', font=('Helvetica Neue', 11, 'bold'))
        style.map('Treeview', background=[('selected', '#bd93f9')], foreground=[('selected', '#282a36')])
        
        # Configure tags for different value types
        self.live_data_tree.tag_configure('good', foreground='#4CAF50')
        self.live_data_tree.tag_configure('fair', foreground='#FF9800')
        self.live_data_tree.tag_configure('poor', foreground='#f44336')
        self.live_data_tree.tag_configure('excellent', foreground='#2E7D32')  # Changed from blue to dark green
        self.live_data_tree.tag_configure('normal', foreground='#000000')
        
        self.live_data_tree.pack(side='left', fill='both', expand=True)
        
        # Scrollbar
        live_data_scrollbar = ttk.Scrollbar(live_data_frame, orient='vertical', command=self.live_data_tree.yview)
        self.live_data_tree.configure(yscrollcommand=live_data_scrollbar.set)
        live_data_scrollbar.pack(side='right', fill='y')
        
        self.log_message("Application started. Ready to export Modbus data.")
        if not MODBUS_AVAILABLE:
            self.log_message("WARNING: pymodbus not installed. Using simulation mode.")

    def open_log_window(self):
        """Open a separate window for the activity log"""
        if self.log_window is not None and self.log_window.winfo_exists():
            # Window already exists, just bring it to front
            self.log_window.lift()
            self.log_window.focus_force()
            return
        
        # Create new log window
        self.log_window = tk.Toplevel(self.root)
        self.log_window.title("Activity Log")
        self.log_window.geometry("800x600")
        self.log_window.configure(bg='#282a36')
        
        # Make window resizable
        self.log_window.resizable(True, True)
        self.log_window.minsize(600, 400)
        
        # Header
        header_frame = tk.Frame(self.log_window, bg='#44475a', relief='flat')
        header_frame.pack(fill='x', pady=(0, 10))
        
        title_label = tk.Label(header_frame, text="Activity Log", 
                              font=("Helvetica Neue", 16, "bold"), 
                              bg='#44475a', fg='#f8f8f2')
        title_label.pack(pady=15)
        
        # Log container
        log_container = tk.Frame(self.log_window, bg='#282a36')
        log_container.pack(fill='both', expand=True, padx=15, pady=(0, 15))
        
        # Log text widget with scrollbar
        self.log_text = tk.Text(log_container, 
                               bg='#282a36', fg='#50fa7b', 
                               font=("SF Mono", 10),
                               relief='flat', bd=0,
                               highlightbackground='#6272a4',
                               highlightcolor='#50fa7b',
                               highlightthickness=1)
        
        log_scrollbar = tk.Scrollbar(log_container, orient='vertical',
                                    command=self.log_text.yview)
        self.log_text.configure(yscrollcommand=log_scrollbar.set)
        
        self.log_text.pack(side='left', fill='both', expand=True)
        log_scrollbar.pack(side='right', fill='y')
        
        # Clear button
        button_frame = tk.Frame(self.log_window, bg='#282a36')
        button_frame.pack(fill='x', padx=15, pady=(0, 15))
        
        clear_btn = tk.Button(button_frame, text="Clear Log",
                             command=self.clear_log,
                             bg='#ff5555', fg='#2d2d2d',
                             font=("Helvetica Neue", 11, "bold"),
                             width=15, height=1, relief='flat', bd=0,
                             activebackground='#ff6e6e',
                             activeforeground='#2d2d2d')
        clear_btn.pack(side='right')
        
        # Handle window closing
        def on_log_window_close():
            self.log_window.destroy()
            self.log_window = None
            self.log_text = None
        
        self.log_window.protocol("WM_DELETE_WINDOW", on_log_window_close)
        
        # Center the window
        self.log_window.update_idletasks()
        x = (self.log_window.winfo_screenwidth() // 2) - (400)
        y = (self.log_window.winfo_screenheight() // 2) - (300)
        self.log_window.geometry(f"800x600+{x}+{y}")
    
    def clear_log(self):
        """Clear the activity log"""
        if self.log_text:
            self.log_text.delete(1.0, tk.END)

    def log_message(self, message):
        """Add a timestamped message to the log"""
        timestamp = datetime.now().strftime("%H:%M:%S")
        log_entry = f"[{timestamp}] {message}\n"
        
        # Print to console
        print(log_entry.strip())
        
        # Add to GUI log if window exists
        if self.log_text:
            self.log_text.insert(tk.END, log_entry)
            self.log_text.see(tk.END)
            if self.log_window and self.log_window.winfo_exists():
                self.log_window.update_idletasks()

    def update_status(self, message, color='#4CAF50'):
        """Update the status label"""
        self.status_label.config(text=message, fg=color)
        self.root.update_idletasks()

    def test_ip(self):
        """Test the IP address connectivity"""
        ip = self.ip_entry.get().strip()
        if not ip:
            messagebox.showerror("Error", "Please enter an IP address")
            return
        
        self.log_message(f"Testing IP address: {ip}")
        self.update_status("Testing connection...", '#FF9800')
        
        # Simulate connection test
        threading.Thread(target=self._test_ip_thread, args=(ip,), daemon=True).start()

    def _test_ip_thread(self, ip):
        """Test IP connection in a separate thread"""
        try:
            if MODBUS_AVAILABLE:
                client = ModbusClient(ip, port=502)
                if client.connect():
                    self.log_message(f"✓ Successfully connected to {ip}")
                    self.update_status("Connection successful", '#4CAF50')
                    self.last_connection_test = True
                    client.close()
                else:
                    self.log_message(f"✗ Failed to connect to {ip}")
                    self.update_status("Connection failed", '#f44336')
                    self.last_connection_test = False
            else:
                # Simulate test
                time.sleep(2)
                self.log_message(f"✓ IP test completed for {ip} (simulation mode)")
                self.update_status("Test completed (simulation)", '#4CAF50')
                self.last_connection_test = True
        except Exception as e:
            self.log_message(f"✗ Error testing IP {ip}: {str(e)}")
            self.update_status("Connection error", '#f44336')
            self.last_connection_test = False
        finally:
            # Update live diagnostics button state based on connection test result
            self.update_live_diagnostics_button()

    def start_export(self):
        """Start the data export process"""
        if self.is_running:
            return
        
        ip = self.ip_entry.get().strip()
        if not ip:
            messagebox.showerror("Error", "Please enter an IP address")
            return
        
        if not self.csv_var.get() and not self.excel_var.get():
            messagebox.showerror("Error", "Please select at least one export format")
            return
        
        self.is_running = True
        self.start_btn.config(state='disabled')
        
        self.log_message(f"Starting export from {ip}")
        self.update_status("Exporting data...", '#FF9800')
        
        # Start export in separate thread
        self.export_thread = threading.Thread(target=self._export_data, args=(ip,), daemon=True)
        self.export_thread.start()

    def stop_export(self):
        """Stop the data export process"""
        if not self.is_running:
            return
        
        self.is_running = False
        self.start_btn.config(state='normal')
        
        self.log_message("Export stopped by user")
        self.update_status("Export stopped", '#FF9800')

    def _export_data(self, ip):
        """Export data using the original collect_data function"""
        try:
            if MODBUS_AVAILABLE:
                # Use original collect_data function
                data = collect_data(ip, self)
                if data:
                    self.log_message(f"Collected {len(data)} device records. Saving files...")
                    
                    # Get base filename once for all exports
                    base_file = self._get_base_filename()
                    if not base_file:
                        self.log_message("Export cancelled by user")
                        return
                    
                    # Check if sensor pairing sheet is requested
                    if self.sensor_pairing_var.get():
                        self._generate_sensor_pairing_sheet(data, base_file)
                    # Always perform normal export if CSV/Excel is selected
                    self._save_original_data_with_base(data, base_file)
                    
                    self.log_message("Export completed successfully!")
                    self.update_status("Export completed", '#4CAF50')
                else:
                    self.log_message("No data collected or connection failed")
                    self.update_status("Export failed", '#f44336')
            else:
                # Simulation mode
                self.log_message("Running in simulation mode...")
                data = []
                for i in range(3):  # Simulate 3 devices
                    if not self.is_running:
                        break
                    
                    device_data = {
                        "DeviceID": 100 + i,
                        "DeviceType": "CL110" if i % 2 == 0 else "TH110",
                        "RFID": f"ABCD{i:04d}",
                        "SerialNumber": f"SN{i:06d}",
                    }
                    data.append(device_data)
                    self.log_message(f"Simulated device {i+1}: {device_data}")
                    time.sleep(1)
                
                if data and self.is_running:
                    self.log_message(f"Collected {len(data)} simulated device records. Saving files...")
                    self._save_original_data(data)
                    self.log_message("Export completed successfully!")
                    self.update_status("Export completed", '#4CAF50')
        
        except Exception as e:
            self.log_message(f"Export error: {str(e)}")
            self.update_status("Export error", '#f44336')
        finally:
            if self.is_running:
                self.is_running = False
                self.start_btn.config(state='normal')

    def flatten_diagnostics(self, data):
        """Flatten the enhanced diagnostics into individual fields for export"""
        # Define the order of common diagnostic fields
        common_fields = [
            "Battery Voltage", "RF Communication Validity", "Communication Status", 
            "Gateway PER", "RSSI", "LQI", "PER Max", "RSSI Min", "LQI Min", "Signal Quality"
        ]
        
        # Define device-specific fields
        device_specific_fields = {
            "HeatTag": ["HeatTag Alarm Type", "HeatTag Alarm Level", "HeatTag Operation Mode"]
        }
        
        # Collect all headers from all devices
        all_headers = set()
        flattened_data = []
        
        for device in data:
            device_type = device.get("DeviceType", "")
            flat_device = {
                "DeviceID": device.get("DeviceID", ""),
                "DeviceType": device_type,
                "RFID": device.get("RFID", ""),
                "SerialNumber": device.get("SerialNumber", ""),
                "DeviceName": device.get("DeviceName", ""),
                "DeviceLabel": device.get("DeviceLabel", "")
            }
            
            diagnostics = device.get("EnhancedDiagnostics", {})
            
            # Add common fields for all devices (if they exist)
            for field in common_fields:
                if field in diagnostics:
                    value = diagnostics[field]
                    # Apply decoders for better readability
                    if field == "Communication Status":
                        value = decode_communication_status(value)
                    elif field == "RF Communication Validity":
                        value = decode_rf_communication_validity(value)
                    flat_device[field] = value
                    all_headers.add(field)
            
            # Add device-specific fields only for the appropriate device types
            for dev_type, fields in device_specific_fields.items():
                if device_type == dev_type:
                    for field in fields:
                        if field in diagnostics:
                            value = diagnostics[field]
                            # Apply HeatTag decoders for better readability
                            if device_type == "HeatTag":
                                if field == "HeatTag Alarm Type":
                                    value = decode_heattag_alarm_type(value)
                                elif field == "HeatTag Alarm Level":
                                    value = decode_heattag_alarm_level(value)
                                elif field == "HeatTag Operation Mode":
                                    value = decode_heattag_operation_mode(value)
                            flat_device[field] = value
                            all_headers.add(field)
            
            flattened_data.append(flat_device)
        
        # Create ordered header list: common fields first, then device-specific fields
        ordered_headers = []
        for field in common_fields:
            if field in all_headers:
                ordered_headers.append(field)
        
        # Add device-specific fields in order
        for dev_type, fields in device_specific_fields.items():
            for field in fields:
                if field in all_headers:
                    ordered_headers.append(field)
        
        return ordered_headers, flattened_data

    def _save_original_data(self, data):
        """Save data in the original format"""
        # Ask user for file location
        base_file = filedialog.asksaveasfilename(
            title="Save Export As",
            defaultextension="",
            filetypes=[("All Files", "*.*")]
        )
        
        if not base_file:
            self.log_message("Export cancelled by user")
            return
        
        # Save as CSV
        if self.csv_var.get():
            filename = base_file + ".csv"
            header_extras, flattened_data = self.flatten_diagnostics(data)
            fieldnames = ["DeviceID", "DeviceType", "RFID", "SerialNumber", "DeviceName", "DeviceLabel"] + header_extras
            with open(filename, "w", newline="") as f:
                writer = csv.DictWriter(f, fieldnames=fieldnames)
                writer.writeheader()
                for row in flattened_data:
                    writer.writerow({k: row.get(k, "") for k in fieldnames})
            self.log_message(f"✓ CSV-Datei gespeichert: {filename}")
        
        # Save as Excel
        if self.excel_var.get() and EXCEL_AVAILABLE:
            filename = base_file + ".xlsx"
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Modbus Export"
            header_extras, flattened_data = self.flatten_diagnostics(data)
            headers = ["DeviceID", "DeviceType", "RFID", "SerialNumber", "DeviceName", "DeviceLabel"] + header_extras
            
            # Add headers
            ws.append(headers)
            
            # Add data rows
            for row in flattened_data:
                ws.append([row.get(h, "") for h in headers])
            
            # Apply conditional formatting for Signal Quality if present
            if "Signal Quality" in headers:
                signal_quality_col = headers.index("Signal Quality") + 1  # Excel columns are 1-indexed
                signal_quality_col_letter = openpyxl.utils.get_column_letter(signal_quality_col)
                
                # Define color fills for different signal quality levels
                from openpyxl.styles import PatternFill
                excellent_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")  # Green
                good_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")      # Light Green
                fair_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")      # Yellow
                weak_fill = PatternFill(start_color="FF6600", end_color="FF6600", fill_type="solid")      # Orange
                very_weak_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid") # Red
                unknown_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")   # Gray
                
                # Apply formatting to each cell in the Signal Quality column
                for row_num in range(2, len(flattened_data) + 2):  # Start from row 2 (after header)
                    cell = ws[f"{signal_quality_col_letter}{row_num}"]
                    signal_quality_value = str(cell.value).strip() if cell.value else ""
                    
                    if signal_quality_value == "Excellent":
                        cell.fill = excellent_fill
                    elif signal_quality_value == "Good":
                        cell.fill = good_fill
                    elif signal_quality_value == "Fair":
                        cell.fill = fair_fill
                    elif signal_quality_value == "Weak":
                        cell.fill = weak_fill
                    elif signal_quality_value == "Very Weak":
                        cell.fill = very_weak_fill
                    elif signal_quality_value == "Unknown" or signal_quality_value == "":
                        cell.fill = unknown_fill
            
            # Apply conditional formatting for RSSI if present
            if "RSSI" in headers:
                rssi_col = headers.index("RSSI") + 1  # Excel columns are 1-indexed
                rssi_col_letter = openpyxl.utils.get_column_letter(rssi_col)
                
                # Define color fills for different RSSI power levels
                from openpyxl.styles import PatternFill
                rssi_good_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")    # Green (0 to -65 dBm)
                rssi_average_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid") # Yellow (-65 to -75 dBm)
                rssi_poor_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")    # Red (< -75 dBm)
                rssi_unknown_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid") # Gray (Unknown/NaN)
                
                # Apply formatting to each cell in the RSSI column
                for row_num in range(2, len(flattened_data) + 2):  # Start from row 2 (after header)
                    cell = ws[f"{rssi_col_letter}{row_num}"]
                    rssi_value = str(cell.value).strip() if cell.value else ""
                    
                    try:
                        if rssi_value and rssi_value.lower() != 'nan':
                            rssi_float = float(rssi_value)
                            if rssi_float >= -65:  # 0 to -65 dBm = Good
                                cell.fill = rssi_good_fill
                            elif rssi_float >= -75:  # -65 to -75 dBm = Average
                                cell.fill = rssi_average_fill
                            else:  # < -75 dBm = Poor
                                cell.fill = rssi_poor_fill
                        else:
                            cell.fill = rssi_unknown_fill  # NaN or empty values
                    except (ValueError, TypeError):
                        cell.fill = rssi_unknown_fill  # Invalid values
            
            wb.save(filename)
            self.log_message(f"✓ Excel-Datei gespeichert: {filename}")

    def _get_base_filename(self):
        """Get base filename for all exports"""
        return filedialog.asksaveasfilename(
            title="Save Export As",
            defaultextension="",
            filetypes=[("All Files", "*.*")]
        )

    def _save_original_data_with_base(self, data, base_file):
        """Save data using the provided base filename"""
        # Add enhanced diagnostics suffix if enabled
        diagnostics_suffix = "_ED" if self.enhanced_diagnostics_var.get() else ""
        
        # Save as CSV
        if self.csv_var.get():
            filename = base_file + diagnostics_suffix + ".csv"
            header_extras, flattened_data = self.flatten_diagnostics(data)
            fieldnames = ["DeviceID", "DeviceType", "RFID", "SerialNumber", "DeviceName", "DeviceLabel"] + header_extras
            with open(filename, "w", newline="") as f:
                writer = csv.DictWriter(f, fieldnames=fieldnames)
                writer.writeheader()
                for row in flattened_data:
                    writer.writerow({k: row.get(k, "") for k in fieldnames})
            self.log_message(f"✓ CSV-Datei gespeichert: {filename}")
        
        # Save as Excel
        if self.excel_var.get() and EXCEL_AVAILABLE:
            # Check if the base filename already ends with .xlsx to avoid double extension
            base_with_suffix = base_file + diagnostics_suffix
            if base_with_suffix.endswith('.xlsx'):
                filename = base_with_suffix
            else:
                filename = base_with_suffix + ".xlsx"
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Modbus Export"
            header_extras, flattened_data = self.flatten_diagnostics(data)
            headers = ["DeviceID", "DeviceType", "RFID", "SerialNumber", "DeviceName", "DeviceLabel"] + header_extras
            
            # Add headers
            ws.append(headers)
            
            # Add data rows
            for row in flattened_data:
                ws.append([row.get(h, "") for h in headers])
            
            # Apply conditional formatting for Signal Quality if present
            if "Signal Quality" in headers:
                signal_quality_col = headers.index("Signal Quality") + 1  # Excel columns are 1-indexed
                signal_quality_col_letter = openpyxl.utils.get_column_letter(signal_quality_col)
                
                # Define color fills for different signal quality levels
                from openpyxl.styles import PatternFill
                excellent_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")  # Green
                good_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")      # Light Green
                fair_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")      # Yellow
                weak_fill = PatternFill(start_color="FF6600", end_color="FF6600", fill_type="solid")      # Orange
                very_weak_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid") # Red
                unknown_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")   # Gray
                
                # Apply formatting to each cell in the Signal Quality column
                for row_num in range(2, len(flattened_data) + 2):  # Start from row 2 (after header)
                    cell = ws[f"{signal_quality_col_letter}{row_num}"]
                    signal_quality_value = str(cell.value).strip() if cell.value else ""
                    
                    if signal_quality_value == "Excellent":
                        cell.fill = excellent_fill
                    elif signal_quality_value == "Good":
                        cell.fill = good_fill
                    elif signal_quality_value == "Fair":
                        cell.fill = fair_fill
                    elif signal_quality_value == "Weak":
                        cell.fill = weak_fill
                    elif signal_quality_value == "Very Weak":
                        cell.fill = very_weak_fill
                    elif signal_quality_value == "Unknown" or signal_quality_value == "":
                        cell.fill = unknown_fill
            
            # Apply conditional formatting for RSSI if present
            if "RSSI" in headers:
                rssi_col = headers.index("RSSI") + 1  # Excel columns are 1-indexed
                rssi_col_letter = openpyxl.utils.get_column_letter(rssi_col)
                
                # Define color fills for different RSSI power levels
                from openpyxl.styles import PatternFill
                rssi_good_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")    # Green (0 to -65 dBm)
                rssi_average_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid") # Yellow (-65 to -75 dBm)
                rssi_poor_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")    # Red (< -75 dBm)
                rssi_unknown_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid") # Gray (Unknown/NaN)
                
                # Apply formatting to each cell in the RSSI column
                for row_num in range(2, len(flattened_data) + 2):  # Start from row 2 (after header)
                    cell = ws[f"{rssi_col_letter}{row_num}"]
                    rssi_value = str(cell.value).strip() if cell.value else ""
                    
                    try:
                        if rssi_value and rssi_value.lower() != 'nan':
                            rssi_float = float(rssi_value)
                            if rssi_float >= -65:  # 0 to -65 dBm = Good
                                cell.fill = rssi_good_fill
                            elif rssi_float >= -75:  # -65 to -75 dBm = Average
                                cell.fill = rssi_average_fill
                            else:  # < -75 dBm = Poor
                                cell.fill = rssi_poor_fill
                        else:
                            cell.fill = rssi_unknown_fill  # NaN or empty values
                    except (ValueError, TypeError):
                        cell.fill = rssi_unknown_fill  # Invalid values
            
            wb.save(filename)
            self.log_message(f"✓ Excel-Datei gespeichert: {filename}")

    def _generate_sensor_pairing_sheet(self, data, base_file):
        """Generate an Excel sensor pairing sheet by merging Modbus data with JSON configuration"""
        json_file = filedialog.askopenfilename(
            title="Select JSON Configuration File",
            filetypes=[("JSON Files", "*.json"), ("All Files", "*.*")]
        )
        
        if not json_file:
            self.log_message("Sensor pairing generation canceled by user")
            return
        
        try:
            self.log_message(f"Loading JSON configuration from {json_file}")
            
            # Load JSON configuration
            with open(json_file, 'r') as json_f:
                json_data = json.load(json_f)
            
            # Get sensors from JSON data
            sensors = json_data.get('sensors', [])
            self.log_message(f"Found {len(sensors)} sensors in configuration")
            
            # Create a mapping from device ID to device data (since RFID formats might differ)
            device_map = {}
            for device in data:
                device_id = str(device.get('DeviceID', ''))
                rfid = device.get('RFID', '')
                if device_id:
                    device_map[device_id] = device
                if rfid:
                    device_map[rfid] = device
            
            # Use the provided base filename and append _SPS
            output_file = f"{base_file}_SPS.xlsx"
            
            # Create Excel workbook
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Sensor Pairing Sheet"
            
            # Updated headers with requested order (removed duplicate Description field)
            headers = [
                "Sensor ID", "RFID", "Serial Number", "Device Type", "Device Name", "Device Label",
                "Equipement", "Sensor Position", "Measured Point", "Cubicle ID", "Cubicle Type",
                "Feeder ID", "Circuit Breaker ID", "Drawer ID"
            ]
            
            # Add headers to worksheet
            ws.append(headers)
            
            # Process each sensor from JSON
            for sensor in sensors:
                sensor_id = sensor.get('slaveId', '')
                rfid = sensor.get('deviceAddress', '')
                # Extract additional attributes from JSON
                equipement = sensor.get('Equipement', '')
                sensor_position = sensor.get('SensorPosition', '')
                measured_point = sensor.get('MeasuredPoint', '')
                cubicle_id = sensor.get('CubicleId', '')
                cubicle_type = sensor.get('CubicleType', '')
                feeder_id = sensor.get('FeederId', '')
                circuit_breaker_id = sensor.get('CircuitBreakerId', '')
                drawer_id = sensor.get('DrawerId', '')
                
                # Find matching device data - try multiple matching strategies
                device_data = device_map.get(rfid, {})
                if not device_data:
                    # Try matching by slaveId (DeviceID)
                    device_data = device_map.get(sensor_id, {})
                
                # Extract device information
                device_type = device_data.get('DeviceType', 'Not Found')
                serial_number = device_data.get('SerialNumber', 'Not Found')
                device_name = device_data.get('DeviceName', 'Not Found')
                device_label = device_data.get('DeviceLabel', 'Not Found')
                
                # Create row data without enhanced diagnostics
                row_data = [
                    sensor_id, rfid, serial_number, device_type, device_name, device_label,
                    equipement, sensor_position, measured_point, cubicle_id, cubicle_type,
                    feeder_id, circuit_breaker_id, drawer_id
                ]
                
                ws.append(row_data)
            
            # No conditional formatting needed for sensor pairing sheet
            
            # Save the workbook
            wb.save(output_file)
            self.log_message(f"✓ Sensor pairing sheet saved: {output_file}")
            
        except Exception as e:
            self.log_message(f"Error generating sensor pairing sheet: {str(e)}")

    def update_live_diagnostics_button(self):
        """Update the live diagnostics button state based on connection status"""
        # Check if the last connection test was successful
        if self.last_connection_test or not MODBUS_AVAILABLE:
            self.live_diag_btn.config(state='normal')
            # Keep status icon red when not active
            self.status_icon.config(fg='#ff5555')
        else:
            self.live_diag_btn.config(state='disabled')
            # Keep status icon red when connection required
            self.status_icon.config(fg='#ff5555')

    def update_live_diagnostics_table(self, live_data=None):
        """Update the live diagnostics table with data or clear it"""
        # Clear existing data
        for item in self.live_data_tree.get_children():
            self.live_data_tree.delete(item)
        
        if live_data:
            # Update timestamp
            current_time = datetime.now().strftime("%H:%M:%S")
            self.last_update_label.config(text=f"Last Update: {current_time}")
            
            # Add new data rows
            for device in live_data:
                device_id = device.get("DeviceID", "Unknown")
                device_type = device.get("DeviceType", "Unknown")
                device_name = device.get("DeviceName", "Unknown")
                diagnostics = device.get("Diagnostics", {})
                
                # Extract values for table columns
                rf_comm = decode_rf_communication_validity(diagnostics.get("RF Communication Validity", "N/A"))
                comm_status = decode_communication_status(diagnostics.get("Communication Status", "N/A"))
                signal_quality = diagnostics.get("Signal Quality", "N/A")
                rssi = diagnostics.get("RSSI", "N/A")
                lqi = diagnostics.get("LQI", "N/A")
                gateway_per = diagnostics.get("Gateway PER", "N/A")
                battery = diagnostics.get("Battery Voltage", "N/A")
                
                # Determine row color based on signal quality
                row_tag = 'normal'
                if signal_quality == "Excellent":
                    row_tag = 'excellent'
                elif signal_quality == "Good":
                    row_tag = 'good'
                elif signal_quality == "Fair":
                    row_tag = 'fair'
                elif signal_quality == "Weak":
                    row_tag = 'poor'
                
                # Prepare data for all columns
                all_data = {
                    "DeviceID": device_id,
                    "DeviceType": device_type,
                    "RFID": device.get("RFID", "Unknown"),
                    "SerialNumber": device.get("SerialNumber", "Unknown"),
                    "DeviceName": device_name,
                    "RFCommunication": rf_comm,
                    "CommStatus": comm_status,
                    "SignalQuality": signal_quality,
                    "RSSI": rssi,
                    "LQI": lqi,
                    "GatewayPER": gateway_per,
                    "Battery": battery
                }
                
                # Get visible columns (always visible + optional selected columns)
                visible_columns = self.always_visible_columns + [col for col in self.optional_columns if self.column_visibility[col].get()]
                
                # Create values list for only visible columns
                values = [all_data.get(col, "") for col in visible_columns]
                
                # Add row to table with color tag
                self.live_data_tree.insert("", "end", values=values, tags=(row_tag,))
            
            # Auto-adjust column widths based on content
            self._auto_adjust_column_widths()
        else:
            # Clear timestamp when no data
            self.last_update_label.config(text="Last Update: Never")
        
        self.root.update_idletasks()

    def toggle_live_diagnostics(self):
        """Toggle live diagnostics on/off"""
        if self.live_diagnostics_enabled:
            self.stop_live_diagnostics()
        else:
            self.start_live_diagnostics()

    def start_live_diagnostics(self):
        """Start live diagnostics monitoring"""
        if self.live_diagnostics_enabled:
            return
        
        ip = self.ip_entry.get().strip()
        if not ip:
            messagebox.showerror("Error", "Please enter an IP address")
            return
        
        self.live_diagnostics_enabled = True
        self.live_diag_btn.config(text="Stop Live Diagnostics", bg='#f44336')
        # Update status icon to green when active
        self.status_icon.config(fg='#50fa7b')
        
        self.log_message("Starting live diagnostics monitoring...")
        self.update_live_diagnostics_table()
        
        # Start live diagnostics in separate thread
        self.live_diagnostics_thread = threading.Thread(target=self._live_diagnostics_worker, args=(ip,), daemon=True)
        self.live_diagnostics_thread.start()

    def stop_live_diagnostics(self):
        """Stop live diagnostics monitoring"""
        if not self.live_diagnostics_enabled:
            return
        
        self.live_diagnostics_enabled = False
        self.live_diag_btn.config(text="Start Live Diagnostics", bg='#4CAF50')
        # Update status icon to red when stopped
        self.status_icon.config(fg='#ff5555')
        
        self.log_message("Live diagnostics monitoring stopped")
        self.update_live_diagnostics_table()

    def _live_diagnostics_worker(self, ip):
        """Worker thread for live diagnostics monitoring"""
        try:
            while self.live_diagnostics_enabled:
                if MODBUS_AVAILABLE:
                    # Collect live data
                    live_data = self._collect_live_diagnostics_data(ip)
                    if live_data:
                        self.update_live_diagnostics_table(live_data)
                    else:
                        self.update_live_diagnostics_table()
                else:
                    # Simulation mode
                    simulated_data = self._simulate_live_diagnostics_data()
                    self.update_live_diagnostics_table(simulated_data)
                
                # Wait for 30 seconds before next update
                for i in range(30):
                    if not self.live_diagnostics_enabled:
                        break
                    time.sleep(1)
                    
        except Exception as e:
            self.log_message(f"Live diagnostics error: {str(e)}")
            self.update_live_diagnostics_table()
            self.stop_live_diagnostics()

    def _collect_live_diagnostics_data(self, ip):
        """Collect live diagnostics data from the device"""
        try:
            client = ModbusClient(ip)
            if not client.connect():
                return None
            
            # Get device IDs
            device_ids = get_device_ids(client)
            if not device_ids:
                client.close()
                return None
            
            live_data = []
            for device_id in device_ids:
                # Get device type first
                ref_regs = read_registers(client, device_id, 31060, 16)
                ref = decode_ascii_cached(ref_regs) if ref_regs else ""
                
                device_type = ""
                if ref == "EMS59443":
                    device_type = "CL110"
                elif ref == "EMS59440":
                    device_type = "TH110"
                elif ref == "SMT10020":
                    device_type = "HeatTag"
                else:
                    device_type = "Unknown"
                
                # Get device name
                device_name_regs = read_registers(client, device_id, 31000, 10)
                device_name = decode_ascii_cached(device_name_regs) if device_name_regs else "Unknown"
                
                # Get RFID
                rfid_regs = read_registers(client, device_id, 31026, 6)
                rfid = ""
                if rfid_regs:
                    hex_str = "".join(f"{reg:04X}" for reg in rfid_regs if reg > 0)
                    rfid = hex_str[:8]
                
                # Get Serial Number
                sn_regs = read_registers(client, device_id, 31088, 10)
                serial_number = decode_ascii_cached(sn_regs) if sn_regs else "Unknown"
                
                # Get enhanced diagnostics
                diagnostics = read_enhanced_diagnostics(client, device_id, device_type)
                
                device_data = {
                    "DeviceID": device_id,
                    "DeviceType": device_type,
                    "DeviceName": device_name,
                    "RFID": rfid,
                    "SerialNumber": serial_number,
                    "Diagnostics": diagnostics
                }
                
                live_data.append(device_data)
            
            client.close()
            return live_data
            
        except Exception as e:
            self.log_message(f"Error collecting live diagnostics data: {str(e)}")
            return None

    def _simulate_live_diagnostics_data(self):
        """Simulate live diagnostics data for demo purposes"""
        import random
        
        # Simulate data for 3 devices
        simulated_data = []
        for i in range(3):
            device_type = "CL110" if i % 2 == 0 else "TH110"
            diagnostics = {
                "RF Communication Validity": 1,
                "Communication Status": 1,
                "Gateway PER": round(random.uniform(5.0, 25.0), 2),
                "RSSI": round(random.uniform(-85.0, -45.0), 2),
                "LQI": random.randint(40, 80),
                "Signal Quality": random.choice(["Good", "Fair", "Excellent"])
            }
            
            if device_type == "CL110":
                diagnostics["Battery Voltage"] = round(random.uniform(3.0, 3.6), 2)
            
            device_data = {
                "DeviceID": 100 + i,
                "DeviceType": device_type,
                "DeviceName": f"Device_{i+1}",
                "RFID": f"ABCD{i:04d}",
                "SerialNumber": f"SN{i:06d}",
                "Diagnostics": diagnostics
            }
            
            simulated_data.append(device_data)
        
        return simulated_data

    def _auto_adjust_column_widths(self):
        """Auto-adjust column widths based on content with improved calculations"""
        # Get list of visible columns (always visible + optional selected columns)
        visible_columns = self.always_visible_columns + [col for col in self.optional_columns if self.column_visibility[col].get()]
        
        # Skip auto-resize for columns that should have fixed widths
        fixed_width_columns = ["DeviceType", "RFID"]
        
        for col in visible_columns:
            # Skip auto-resize for fixed width columns
            if col in fixed_width_columns:
                continue
            # Get the header text width (headers are bold, so need more space)
            header_text = self.live_data_tree.heading(col, 'text')
            header_width = len(header_text) * 12  # Bold headers need more space
            
            # Find the maximum width for this column
            max_content_width = 0
            for item in self.live_data_tree.get_children():
                visible_col_index = visible_columns.index(col)
                try:
                    value = str(self.live_data_tree.item(item, 'values')[visible_col_index])
                    # Better content width calculation - account for different character widths
                    content_width = len(value) * 10  # Regular text width
                    max_content_width = max(max_content_width, content_width)
                except IndexError:
                    continue
            
            # Take the maximum of header and content widths
            calculated_width = max(header_width, max_content_width)
            
            # Set column-specific minimum and maximum widths with fixed widths for Type and RFID
            column_limits = {
                "DeviceID": {"min": 60, "max": 120},
                "DeviceType": {"min": 110, "max": 110},  # Fixed width for "HeatTag" (7 chars)
                "RFID": {"min": 120, "max": 120},  # Fixed width for 8 chars (increased)
                "SerialNumber": {"min": 120, "max": 180},
                "DeviceName": {"min": 120, "max": 300},
                "RFCommunication": {"min": 80, "max": 140},
                "CommStatus": {"min": 100, "max": 180},
                "SignalQuality": {"min": 100, "max": 150},
                "RSSI": {"min": 80, "max": 140},
                "LQI": {"min": 60, "max": 100},
                "GatewayPER": {"min": 80, "max": 140},
                "Battery": {"min": 80, "max": 140}
            }
            
            limits = column_limits.get(col, {"min": 60, "max": 200})
            min_width = limits["min"]
            max_width_limit = limits["max"]
            
            # Add generous padding and apply limits
            final_width = calculated_width + 30  # More generous padding
            final_width = min(final_width, max_width_limit)
            final_width = max(min_width, final_width)
            
            # Apply the width
            self.live_data_tree.column(col, width=final_width)

    def update_column_visibility(self):
        """Update which columns are visible in the live diagnostics table"""
        # Prevent GUI updates during column reconfiguration
        self.root.update_idletasks()
        
        # Get list of visible columns (always visible + optional selected columns)
        visible_columns = self.always_visible_columns + [col for col in self.optional_columns if self.column_visibility[col].get()]
        
        # Update the tree view to show only visible columns
        self.live_data_tree.config(columns=visible_columns)
        
        # Reset column headings for visible columns
        column_display_names = {
            "DeviceID": "ID",
            "DeviceType": "Type",
            "RFID": "RFID",
            "SerialNumber": "Serial Number",
            "DeviceName": "Name",
            "RFCommunication": "RF Com.",
            "CommStatus": "Com. Status",
            "SignalQuality": "Signal Quality",
            "RSSI": "RSSI (dBm)",
            "LQI": "LQI",
            "GatewayPER": "Gateway PER",
            "Battery": "Battery (V)"
        }
        
        for col in visible_columns:
            self.live_data_tree.heading(col, text=column_display_names.get(col, col))
        
        # Clear existing data and refresh if live diagnostics is running
        for item in self.live_data_tree.get_children():
            self.live_data_tree.delete(item)
        
        # If live diagnostics is running, trigger a refresh
        if self.live_diagnostics_enabled:
            # The data will be refreshed on the next update cycle
            pass
        
        # Auto-adjust column widths without affecting overall layout
        self._auto_adjust_column_widths()
        
        # Ensure the main window geometry remains stable
        self.root.update_idletasks()

    def on_closing(self):
        """Handle application closing"""
        if self.is_running or self.live_diagnostics_enabled:
            message = "Export is running. Do you want to stop and quit?"
            if self.live_diagnostics_enabled:
                message = "Live diagnostics is running. Do you want to stop and quit?"
            if self.is_running and self.live_diagnostics_enabled:
                message = "Export and live diagnostics are running. Do you want to stop and quit?"
            
            if messagebox.askokcancel("Quit", message):
                if self.is_running:
                    self.stop_export()
                if self.live_diagnostics_enabled:
                    self.stop_live_diagnostics()
                self.root.after(1000, self.root.destroy)  # Give time for cleanup
        else:
            self.root.destroy()

def main():
    """Main application entry point"""
    root = tk.Tk()
    app = ModbusExporterGUI(root)
    
    # Center the window
    root.update_idletasks()
    width = root.winfo_width()
    height = root.winfo_height()
    x = (root.winfo_screenwidth() // 2) - (width // 2)
    y = (root.winfo_screenheight() // 2) - (height // 2)
    root.geometry(f'{width}x{height}+{x}+{y}')
    
    root.mainloop()

if __name__ == "__main__":
    main()

